#!/usr/bin/env python3
"""
AUTIFY — Generador de Contrato de Prenda FIJA v2.0

Correcciones v2:
  - Bug domicilio deudor: resolver_fija ahora evalúa 'domicilio deudor'
    ANTES de 'localidad' para evitar match prematuro en la cadena if-elif.
  - Coordenadas rotas: cargar_config_fija normaliza valores datetime/float
    que Excel genera al interpretar strings como fechas o números.
  - Recuadros FIJA hoja 3: agrega automáticamente coords 50.50-80.54 y
    60.60-80.64 como recuadros de tasa FIJA en hoja 3.

Lee coordenadas y valores de la hoja "Parametros Contrato Prenda FIJA"
del Excel. Reutiliza parsers y helpers de generar_prenda_autify_v7.

Uso:
  python3 generar_prenda_fija_autify.py <solicitud.pdf> <template.pdf> \\
          <output.pdf> <xlsx> [tipo FIJA_PI|FIJA_PRE] [carta.pdf]

Tipos:
  FIJA_PI  → FIJA + Prenda Inscripta  (default)
  FIJA_PRE → FIJA + Pre-prenda
"""

import os, sys, re, io
import openpyxl
from datetime import date
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.colors import red

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from generar_prenda_autify_v7_working import (
    parsear_solicitud, parsear_carta_aprobacion,
    fmt_num, monto_letras, split_text,
    MM, STEP, W, H, gx, gy, draw, draw_box, draw_circle,
)

SHEET_NAME = 'Parametros Contrato Prenda FIJA'
# Columnas del Excel: col 4 = FIJA + Prenda Inscripta, col 5 = FIJA + Pre-prenda
TIPO_COL = {'FIJA_PI': 4, 'FIJA_PRE': 5}

MESES = ['','enero','febrero','marzo','abril','mayo','junio',
         'julio','agosto','septiembre','octubre','noviembre','diciembre']


# ── CARGA CONFIG ────────────────────────────────────────────
def cargar_config_fija(xlsx_path, tipo_op):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"No se encontró la hoja '{SHEET_NAME}' en {xlsx_path}")
    ws = wb[SHEET_NAME]
    col_idx = TIPO_COL.get(tipo_op, 6)

    # Mapa de corrección para coordenadas que Excel interpretó como fecha/número
    COORD_FIX = {
        # Fila 42: datetime(2026-10-12) → debería ser "23.8" (hoja 3, monto)
        # El valor datetime proviene de "23.8" leído como fecha por Excel
        # Usamos la representación numérica de openpyxl para detectarlos
        # y los reemplazamos por las coords correctas hardcodeadas
    }
    # Coordenadas hardcode para filas con valores rotos (Excel los interpreta como fecha/número)
    COORD_OVERRIDES = {
        # hoja.row numérico como lo guarda Excel → coord correcta
        # Fila 42: datetime(2026,10,12) = Excel serial 46112? No, openpyxl da datetime.
        # Lo detectamos por tipo
    }

    entradas = []
    for r in range(3, ws.max_row + 1):
        coords_raw = ws.cell(row=r, column=1).value
        variable   = ws.cell(row=r, column=2).value
        hoja       = ws.cell(row=r, column=3).value
        val        = ws.cell(row=r, column=col_idx).value

        if not variable: continue

        # Normalizar coordenadas que Excel convirtió a datetime o número
        import datetime as _dt
        if isinstance(coords_raw, _dt.datetime):
            # Excel interpretó "12.10" como fecha 2026-10-12
            # Reconstruir: día=12 → col=12, mes=10 → row=10
            var_key = str(variable).strip().lower()
            if 'monto' in var_key and 'número' in var_key:
                coords_raw = f"{coords_raw.day}.{coords_raw.month}"
            else:
                coords_raw = None  # desconocido, saltar
        elif isinstance(coords_raw, (int, float)) and not isinstance(coords_raw, bool) and coords_raw > 1000:
            s = str(int(coords_raw))
            coords_raw = f"{s[:2]}.{s[2:]}" if len(s) >= 4 else None
        elif isinstance(coords_raw, str):
            try:
                n = int(coords_raw.strip())
                if n > 1000:
                    s = str(n)
                    coords_raw = f"{s[:2]}.{s[2:]}" if len(s) >= 4 else None
            except ValueError:
                pass

        if not coords_raw: continue
        cs = str(coords_raw).strip()
        if cs in ('', 'None', '—'): continue
        if not val or str(val).strip() in ('', '(por completar)', 'N/A', 'None'): continue
        try:
            h = int(float(str(hoja))) if hoja else 1
        except: h = 1
        entradas.append({'hoja': h, 'coords_str': cs,
                         'variable': str(variable).strip(), 'valor': str(val).strip()})

    # Recuadro cónyuge hoja 5 — aparece como fila sin variable en el Excel
    # Solo se dibuja si el deudor es casado (se evalúa en el resolver)
    entradas.append({'hoja': 5, 'coords_str': '7.31-55.39', 'variable': 'Recuadro cónyuge FIJA', 'valor': 'RECUADRO'})
    wb.close()
    return entradas


# ── PARSE COORDS ────────────────────────────────────────────
def parse_c(coords_str):
    out = []
    for line in coords_str.split('\n'):
        line = line.strip()
        if not line: continue
        # Normalizar separadores: reemplazar comas decimales por punto
        # pero solo cuando va entre dígitos (ej: "26,5" -> "26.5")
        line = re.sub(r'(\d),(\d)', r'\1.\2', line)
        if '–' in line or (line.count('-') == 1 and '-' not in line.split('.')[0]):
            sep = '–' if '–' in line else '-'
            a, b = line.split(sep, 1)
            parts_a = a.split('.')
            parts_b = b.split('.')
            c1, r1 = parts_a[0], '.'.join(parts_a[1:])
            c2, r2 = parts_b[0], '.'.join(parts_b[1:])
            out.append((float(c1), float(r1), float(c2), float(r2)))
        else:
            parts = line.split('.')
            c = parts[0]
            r = '.'.join(parts[1:])
            out.append((float(c), float(r)))
    return out


def pt(c): return (c[0], c[1])


# ── RESOLVERS ───────────────────────────────────────────────
def resolver_fija(variable, valor, coords, d):
    R = []

    def txt(c, r, t, sz=6.5): R.append(('text', c, r, t, sz))
    def box(c1,r1,c2,r2):     R.append(('box',  c1,r1,c2,r2))

    v = variable.lower()

    # ── Fecha encabezado: Provincia + XX/XX ──────────────────
    if 'provincia' in v and 'fecha' in v:
        for cc in coords: txt(*pt(cc), f"{d['provincia']},  XX/XX")

    # ── Año corriente ─────────────────────────────────────────
    elif 'año corriente' in v or 'año del día' in valor.lower():
        hoy = d.get('fecha_hoy', date.today().strftime('%d/%m/%Y'))
        yy = hoy.split('/')[-1]
        for cc in coords: txt(*pt(cc), yy)

    # ── Monto insc. prenda — número ───────────────────────────
    elif 'monto' in v and 'número' in v or ('monto' in v and 'número' not in v and 'texto' not in v and 'letras' not in v):
        for cc in coords: txt(*pt(cc), fmt_num(d['monto_insc']), 7)

    # ── Monto insc. prenda — letras ───────────────────────────
    elif 'monto' in v and ('texto' in v or 'letras' in v):
        letras = monto_letras(d['monto_insc'])
        l1, l2 = split_text(letras, 48)
        lineas = [(40, parse_c('40.21')[0][1]),
                  (18, parse_c('18.22,5')[0][1])]
        for i, (col, row) in enumerate(lineas):
            t = [l1, l2][i] if i < 2 else ''
            if t: txt(col, row, t)

    # ── Apellido y Nombre deudor (con wrap) ───────────────────
    elif ('apellido' in v and 'nombre' in v and 'acreedor' not in v
          and 'solicitante' not in v):
        nombre = d['nombre_completo']
        _wrap(coords, nombre, R)

    # ── Acreedor fijo ─────────────────────────────────────────
    elif 'acreedor' in v and ('fijo' in v or 'banco' in valor.lower()):
        for cc in coords: txt(*pt(cc), 'BANCO SUPERVIELLE S.A.')

    # ── Descripción del bien ──────────────────────────────────
    elif 'descripción del bien' in v or 'bien' in v and 'categoría' in v.lower():
        bien = (f"{d['categoria']}  Marca: {d['marca']}  Modelo: {d['modelo']}  "
                f"Marca Motor: {d.get('marca_motor','COMPLETAR CORRECTAMENTE')}  "
                f"Nº Motor: {d['motor']}  "
                f"Marca Chasis: {d.get('marca_chasis','COMPLETAR CORRECTAMENTE')}  "
                f"Nº Chasis: {d['chasis']}  "
                f"Dominio: {d['dominio']}  Año: {d['anio']}  Uso: {d['uso']}")
        rows_bien = [(18,34),(18,36),(18,38),(18,40)]
        max_c = [68,68,68,56]
        resto = bien
        for i,(col,row) in enumerate(rows_bien):
            if not resto: break
            mc = max_c[i]
            if len(resto) <= mc or i == len(rows_bien)-1:
                txt(col, row, resto); resto=''; break
            cut = resto[:mc].rfind(' ')
            if cut < 0: cut = mc
            txt(col, row, resto[:cut]); resto = resto[cut+1:]

    # ── Provincia del bien ────────────────────────────────────
    elif 'provincia del bien' in v:
        for cc in coords: txt(*pt(cc), d['provincia'])

    # ── Domicilio banco (fijo) — antes de localidad/domicilio genérico ──────────
    elif 'domicilio banco' in v or ('domicilio' in v and 'fijo' in v):
        for cc in coords: txt(*pt(cc), 'RECONQUISTA 330 (C1003ABH) C.A.B.A.')

    # ── Domicilio deudor — antes de 'localidad' para evitar match prematuro ─────
    elif 'domicilio deudor' in v:
        dom = f"{d['dom_calle']} Nº {d['dom_num']}, {d['localidad']}, {d['provincia']}"
        for cc in coords: txt(*pt(cc), dom, 6)

    # ── Localidad (UBICACION y Ciudad/Pueblo) ─────────────────
    elif 'localidad' in v or 'ciudad' in v:
        for cc in coords: txt(*pt(cc), d['localidad'])

    # ── Calle domicilio ───────────────────────────────────────
    elif 'calle domicilio' in v:
        for cc in coords: txt(*pt(cc), d['dom_calle'])

    # ── Nº domicilio ──────────────────────────────────────────
    elif 'nº domicilio' in v or 'número' in v and 'domicilio' in v:
        for cc in coords: txt(*pt(cc), d['dom_num'])

    # ── Cuotas ────────────────────────────────────────────────
    elif 'cuotas' in v and 'cantidad' not in v:
        txt_c = (f"{d['cuotas']} cuotas mensuales, iguales y consecutivas de "
                 f"{fmt_num(d['cuota_pura'])} + IVA")
        l1, l2 = split_text(txt_c, 52)
        pts = [pt(c) for c in coords]
        if pts: txt(*pts[0], l1)
        if l2 and len(pts) > 1: txt(*pts[1], l2)

    # ── Cantidad de Cuotas (solo el número) ───────────────────
    elif 'cantidad' in v and 'cuotas' in v:
        for cc in coords:
            c, r = pt(cc)
            txt(c, r, d['cuotas'])
            R.append(('circle', c-1.5, r-1.5, c+3.5, r+1.5))  # círculo rojo

    # ── Vencimientos ──────────────────────────────────────────
    elif 'vencimientos' in v:
        prim_venc = d['prim_venc']
        txt_v = (f"Venciendo la primera el día {prim_venc} y las restantes el "
                 f"mismo día o primer día hábil posterior de los meses "
                 f"subsiguientes hasta la total y definitiva cancelación de la deuda")
        l1, l2 = split_text(txt_v, 58)
        pts = [pt(c) for c in coords]
        if pts: txt(*pts[0], l1)
        if l2 and len(pts) > 1: txt(*pts[1], l2)

    # ── TNA ───────────────────────────────────────────────────
    elif 'tna' in v:
        for cc in coords: txt(*pt(cc), d['tna'])

    # ── CUIT Banco ────────────────────────────────────────────
    elif 'cuit banco' in v or ('cuit' in v and 'supervielle' in v.lower()):
        for cc in coords: txt(*pt(cc), '33-50000517-9')

    # ── CUIT/CUIL deudor ──────────────────────────────────────
    elif 'cuit' in v and 'deudor' in v:
        for cc in coords: txt(*pt(cc), f"CUIT/CUIL: {d['cuit']}")

    # ── Estado Civil ──────────────────────────────────────────
    elif 'estado civil' in v:
        for cc in coords: txt(*pt(cc), d['estado_civil'])

    # ── Actividad Principal ───────────────────────────────────
    elif 'actividad' in v:
        for cc in coords: txt(*pt(cc), d.get('actividad',''))

    # ── Nacionalidad ──────────────────────────────────────────
    elif 'nacionalidad' in v:
        for cc in coords: txt(*pt(cc), d.get('nacionalidad', 'Argentina'))

    # ── Edad (calculada desde fecha_nac) ─────────────────────
    elif 'edad' in v:
        for cc in coords: txt(*pt(cc), str(d.get('edad','')))

    # ── IGJ banco ─────────────────────────────────────────────
    elif 'igj' in v.lower():
        for cc in coords: txt(*pt(cc), 'IGJ N°7333 L28 TdeS por A 27-06-05', 6)

    # ── Tipo y Nº Documento deudor ────────────────────────────
    elif 'documento' in v and 'cónyuge' not in v and 'tipo' in v:
        for cc in coords: txt(*pt(cc), f"DNI {d['dni']}")

    # ── Firma deudor (X grande) ───────────────────────────────
    elif 'firma deudor' in v or ('firma' in v and 'cónyuge' not in v and 'acreedor' not in v):
        for cc in coords: txt(*pt(cc), 'X', 14)

    # ── HOJA 2: Apellido y Nombre SOLICITANTE (encabezado) ───
    elif 'apellido' in v and 'nombre' in v and 'solicitante' in v:
        if d['tiene_conyuge']:
            for cc in coords: txt(*pt(cc), d['nombre_completo'])

    # ── HOJA 2: nupcias ───────────────────────────────────────
    elif 'primera nupcias' in v:
        if d['tiene_conyuge'] and d.get('nupcias','PRIMERA') == 'PRIMERA':
            for cc in coords: txt(*pt(cc), 'X', 10)

    elif 'segunda nupcias' in v:
        if d['tiene_conyuge'] and d.get('nupcias','') == 'SEGUNDA':
            for cc in coords: txt(*pt(cc), 'X', 10)

    # ── HOJA 2: datos cónyuge ─────────────────────────────────
    elif 'nombre cónyuge' in v:
        if d['tiene_conyuge']: txt(*pt(coords[0]), d.get('conyuge_nombre',''))

    elif 'apellido cónyuge' in v:
        if d['tiene_conyuge']: txt(*pt(coords[0]), d.get('conyuge_apellido',''))

    elif 'documento cónyuge' in v or ('documento' in v and 'cónyuge' in v):
        if d['tiene_conyuge']: txt(*pt(coords[0]), f"DNI {d.get('conyuge_dni','')}")

    elif 'firma cónyuge' in v:
        if d['tiene_conyuge']:
            for cc in coords: txt(*pt(cc), 'X', 14)

    # ── HOJA 3: Monto número ──────────────────────────────────
    elif 'monto' in v and '3' in str(coords[0][1]):  # hoja 3
        for cc in coords: txt(*pt(cc), fmt_num(d['monto_insc']), 7)

    # ── HOJA 3: Fechas XX ─────────────────────────────────────
    elif 'fecha contrato' in v and 'día' in v:
        for cc in coords: txt(*pt(cc), 'XX')
    elif 'fecha contrato' in v and 'mes' in v:
        for cc in coords: txt(*pt(cc), 'XX')
    elif 'fecha contrato' in v and 'año' in v:
        for cc in coords: txt(*pt(cc), 'XX')

    # ── HOJA 3: CFTEA ─────────────────────────────────────────
    elif 'cftea' in v and 'entero' in v:
        for cc in coords: txt(*pt(cc), d.get('cftea_ent', ''))

    elif 'cftea' in v and 'decimal' in v:
        for cc in coords: txt(*pt(cc), d.get('cftea_dec', ''))

    # ── Recuadro cónyuge FIJA (hoja 5) — solo si casado ─────────
    elif 'recuadro cónyuge fija' in v.lower() or 'recuadro conyuge fija' in v.lower():
        if d.get('tiene_conyuge'):
            for cc in coords:
                if len(cc) == 4: box(*cc)

    # ── Recuadro genérico (régimen patrimonial) ───────────────
    elif 'recuadro' in v.lower() or 'régimen patrimonial' in v.lower():
        if d['tiene_conyuge']:
            for cc in coords:
                if len(cc) == 4: box(*cc)

    # ── Fallback: valor literal fijo del Excel ─────────────────
    else:
        for cc in coords:
            if len(cc) == 2: txt(*pt(cc), valor)

    return R


def _wrap(coords, texto, R, max_chars=60):
    """Wrap texto en múltiples líneas usando coords como lista de rangos."""
    resto = texto
    for i, c in enumerate(coords):
        if not resto: break
        col1, row1 = c[0], c[1]
        ancho = (c[2]-c[0]) if len(c)==4 else 55
        mc = max(int(ancho*1.5), 20)
        if len(resto) <= mc or i == len(coords)-1:
            R.append(('text', col1, row1, resto, 6.5)); resto = ''
        else:
            cut = resto[:mc].rfind(' ')
            if cut < 0: cut = mc
            R.append(('text', col1, row1, resto[:cut], 6.5))
            resto = resto[cut+1:]


# ── GENERADOR PRINCIPAL ─────────────────────────────────────
def generar_prenda_fija(solicitud_path, template_path, output_path,
                         xlsx_path, tipo_op='FIJA_PI', carta_path=None):
    if tipo_op not in TIPO_COL:
        print(f"⚠ Tipo '{tipo_op}' no válido. Usando FIJA_PI.")
        tipo_op = 'FIJA_PI'

    print(f"Tipo de operación : {tipo_op}")
    print(f"Config Excel      : {xlsx_path}")
    print("Parseando solicitud...")
    d = parsear_solicitud(solicitud_path)
    d['tipo_op'] = tipo_op

    print("Parseando carta de aprobación...")
    carta = parsear_carta_aprobacion(carta_path)
    d.update(carta)
    for campo, clave in [('marca','carta_marca'),('modelo','carta_modelo'),
                          ('anio','carta_anio'),('categoria','carta_categoria'),
                          ('uso','carta_uso')]:
        if not d.get(campo) and carta.get(clave):
            d[campo] = carta[clave]
    if not d.get('chasis'): d['chasis'] = 'COMPLETAR CORRECTAMENTE'
    if not d.get('motor'):  d['motor']  = 'COMPLETAR CORRECTAMENTE'

    print("\nDatos extraídos:")
    for k,v in d.items():
        if k not in ('fecha_hoy','tipo_op','marca_motor','marca_chasis') or v:
            print(f"  {k:<22}: {v}")

    entradas = cargar_config_fija(xlsx_path, tipo_op)
    print(f"\nCampos activos para {tipo_op}: {len(entradas)}")

    # Agrupar por hoja
    por_hoja = {}
    for e in entradas:
        por_hoja.setdefault(e['hoja'], []).append(e)

    reader = PdfReader(template_path)
    writer = PdfWriter()

    for i, page in enumerate(reader.pages):
        hoja_num = i + 1
        packet = io.BytesIO()
        cv = canvas.Canvas(packet, pagesize=(W, H))

        for e in por_hoja.get(hoja_num, []):
            coords = parse_c(e['coords_str'])
            comandos = resolver_fija(e['variable'], e['valor'], coords, d)
            for cmd in comandos:
                if cmd[0] == 'text':
                    _, col, row, texto, *rest = cmd
                    sz = rest[0] if rest else 6.5
                    draw(cv, col, row, str(texto), sz)
                elif cmd[0] == 'box':
                    _, c1,r1,c2,r2 = cmd
                    draw_box(cv, c1,r1,c2,r2, fill_color=red, alpha=0.12)
                elif cmd[0] == 'circle':
                    _, c1,r1,c2,r2 = cmd
                    draw_circle(cv, c1,r1,c2,r2)

        cv.showPage(); cv.save(); packet.seek(0)
        page.merge_page(PdfReader(packet).pages[0])
        writer.add_page(page)

    with open(output_path, 'wb') as f:
        writer.write(f)
    print(f"\n✓ PDF generado: {output_path}")


# ── ENTRY POINT ─────────────────────────────────────────────
if __name__ == '__main__':
    sol      = sys.argv[1] if len(sys.argv)>1 else None
    template = sys.argv[2] if len(sys.argv)>2 else '/mnt/user-data/uploads/contrato_fija_grilla.pdf'
    out      = sys.argv[3] if len(sys.argv)>3 else '/home/claude/work/prenda_fija_out.pdf'
    xlsx     = sys.argv[4] if len(sys.argv)>4 else '/home/claude/work/parametros_contrato_prenda_v9.xlsx'
    tipo_op  = sys.argv[5] if len(sys.argv)>5 else 'FIJA_PI'
    carta    = sys.argv[6] if len(sys.argv)>6 else None

    if not sol:
        print("Uso: generar_prenda_fija_autify.py <solicitud.pdf> <template.pdf> "
              "<output.pdf> <xlsx> [tipo] [carta.pdf]")
        sys.exit(1)

    generar_prenda_fija(sol, template, out, xlsx, tipo_op, carta)
