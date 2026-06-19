#!/usr/bin/env python3
"""
AUTIFY — Generador de Formulario 03 v1.0

Lee coordenadas y valores de la hoja "Parametros 03" del Excel.
Reutiliza parsers y helpers de generar_prenda_autify_v7.

Uso:
  python3 generar_form03_autify.py <solicitud.pdf> <template03.pdf> \
          <output.pdf> <xlsx> [tipo UVA_PI|UVA_PRE|FIJA_PI|FIJA_PRE]

Tipos:
  UVA_PI   → UVA + Prenda Inscripta  (col 4)
  UVA_PRE  → UVA + Pre-prenda        (col 5)
  FIJA_PI  → FIJA + Prenda Inscripta (col 6)
  FIJA_PRE → FIJA + Pre-prenda       (col 7)
"""

import os, sys, re, io
import openpyxl
import datetime as _dt
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.colors import red

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from generar_prenda_autify_v7_working import (
    parsear_solicitud, parsear_carta_aprobacion,
    fmt_num, split_text,
    MM, STEP, W, H, gx, gy, draw, draw_box,
)

SHEET_NAME = 'Parametros 03'
TIPO_COL = {
    'UVA_PI':  4,
    'UVA_PRE': 5,
    'FIJA_PI': 6,
    'FIJA_PRE': 7,
}


def cargar_config_03(xlsx_path, tipo_op):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"No se encontró la hoja '{SHEET_NAME}' en {xlsx_path}")
    ws = wb[SHEET_NAME]
    col_idx = TIPO_COL.get(tipo_op, 4)

    entradas = []
    for r in range(2, ws.max_row + 1):
        coords_raw = ws.cell(row=r, column=1).value
        variable   = ws.cell(row=r, column=2).value
        val        = ws.cell(row=r, column=col_idx).value

        if not variable: continue

        # Normalizar coordenadas que Excel convirtió a float (ej: 27101.0 → "27.101")
        if isinstance(coords_raw, _dt.datetime):
            coords_raw = None
        elif isinstance(coords_raw, float) and coords_raw > 1000:
            s = str(int(coords_raw))
            if len(s) >= 4:
                coords_raw = f"{s[:2]}.{s[2:]}"
            else:
                coords_raw = None

        if not coords_raw: continue
        cs = str(coords_raw).strip()
        if cs in ('', 'None', '—'): continue
        if not val or str(val).strip() in ('', '(por completar)', 'N/A', 'None'): continue

        entradas.append({
            'coords_str': cs,
            'variable': str(variable).strip(),
            'valor': str(val).strip()
        })

    wb.close()
    return entradas


def parse_c(coords_str):
    """Parsea string de coordenadas a lista de tuplas (col, row) o (c1,r1,c2,r2)."""
    out = []
    for line in coords_str.split('\n'):
        line = line.strip()
        if not line: continue
        line = re.sub(r'(\d),(\d)', r'\1.\2', line)
        if '–' in line or (line.count('-') == 1 and not line.startswith('-')):
            sep = '–' if '–' in line else '-'
            a, b = line.split(sep, 1)
            pa = a.split('.'); pb = b.split('.')
            c1, r1 = pa[0], '.'.join(pa[1:])
            c2, r2 = pb[0], '.'.join(pb[1:])
            out.append((float(c1), float(r1), float(c2), float(r2)))
        else:
            parts = line.split('.')
            c = parts[0]; r = '.'.join(parts[1:])
            out.append((float(c), float(r)))
    return out


def pt(c): return (c[0], c[1])


def resolver_03(variable, valor, coords, d):
    R = []
    v = variable.lower().strip()

    def txt(c, r, t, sz=6.5): R.append(('text', c, r, t, sz))
    def box(c1,r1,c2,r2):     R.append(('box',  c1,r1,c2,r2))

    # ── Sección A: fecha, monto, dominio ─────────────────────
    if v == 'dia':
        for cc in coords: txt(*pt(cc), 'XX')
    elif v == 'mes':
        for cc in coords: txt(*pt(cc), 'XX')
    elif v == 'año':
        for cc in coords: txt(*pt(cc), 'XX')
    elif 'monto insc' in v:
        for cc in coords: txt(*pt(cc), fmt_num(d['monto_insc']), 7)
    elif 'dominio' in v and 'garantía' in v or 'dominio' in v:
        for cc in coords: txt(*pt(cc), d['dominio'])

    # ── Sección D: acreedor (fijo) ────────────────────────────
    elif 'cuit acreedor' in v:
        for cc in coords: txt(*pt(cc), '33-50000517-9')
    elif 'apellido y nombre acreedor' in v or ('apellido' in v and 'acreedor' in v):
        # coords es lista de rangos [c1,r1,c2,r2] → usamos solo la primera línea
        # BANCO SUPERVIELLE S.A. cabe en una sola línea
        if coords: txt(coords[0][0], coords[0][1], 'BANCO SUPERVIELLE S.A.')
    elif 'calle acreedor' in v:
        for cc in coords: txt(*pt(cc), 'RECONQUISTA')
    elif 'numero acreedor' in v:
        for cc in coords: txt(*pt(cc), '330')
    elif 'cp acreedor' in v:
        for cc in coords: txt(*pt(cc), 'C1003ABH')
    elif 'localidad acreedor' in v:
        for cc in coords: txt(*pt(cc), 'C.A.B.A.')
    elif 'personeria otorgada' in v:
        for cc in coords: txt(*pt(cc), 'I.G.J')
    elif 'datos de inscripcion' in v:
        for cc in coords: txt(*pt(cc), 'N°7333 L28 TdeS por A 27-06-05', 6)
    elif v == 'dia inscripcion':
        for cc in coords: txt(*pt(cc), '27')
    elif v == 'mes inscripcion':
        for cc in coords: txt(*pt(cc), '06')
    elif v == 'año inscripcion':
        for cc in coords: txt(*pt(cc), '05')

    # ── Sección E: deudor ─────────────────────────────────────
    elif 'cuil deudor' in v:
        for cc in coords: txt(*pt(cc), d['cuit'], 6.5)
    elif 'apellido y nombre deudor' in v or ('apellido' in v and 'deudor' in v):
        # coords tiene hasta 3 rangos (líneas 28, 32, 34 del form03)
        # Calcular ancho en chars de la primera línea: (c2-c1) * ~1.5 chars/unidad ≈ 30 chars
        nombre = d['nombre_completo']
        l1, l2 = split_text(nombre, 26)
        if coords:
            txt(coords[0][0], coords[0][1], l1)
        if l2 and len(coords) > 1:
            txt(coords[1][0], coords[1][1], l2)
    elif 'calle deudor' in v:
        for cc in coords: txt(*pt(cc), d['dom_calle'])
    elif 'numero deudor' in v:
        for cc in coords: txt(*pt(cc), str(d['dom_num']))
    elif 'piso deudor' in v:
        if d.get('dom_piso') and str(d['dom_piso']) not in ('0', ''):
            for cc in coords: txt(*pt(cc), str(d['dom_piso']))
    elif 'dpto deudor' in v:
        if d.get('dom_depto') and str(d['dom_depto']) not in ('', 'None'):
            for cc in coords: txt(*pt(cc), str(d['dom_depto']))
    elif 'cp deudor' in v:
        for cc in coords: txt(*pt(cc), str(d['cod_postal']))
    elif 'localidad deudor' in v:
        for cc in coords: txt(*pt(cc), d['localidad'])
    elif 'provincia deudor' in v:
        for cc in coords: txt(*pt(cc), d['provincia'])
    elif v == 'dni':
        # "Tipo y Nº Documento" → ej: "DNI 40472714"
        tipo_doc = d.get('tipo_doc', 'DNI')
        for cc in coords: txt(*pt(cc), f"{tipo_doc} {d['dni']}")
    elif 'dia nacimiento' in v:
        dd = d['fecha_nac'].split('/')[0] if d.get('fecha_nac') else ''
        for cc in coords: txt(*pt(cc), dd)
    elif 'mes nacimiento' in v:
        mm = d['fecha_nac'].split('/')[1] if d.get('fecha_nac') else ''
        for cc in coords: txt(*pt(cc), mm)
    elif 'año nacimiento' in v:
        aa = d['fecha_nac'].split('/')[-1][-2:] if d.get('fecha_nac') else ''
        for cc in coords: txt(*pt(cc), aa)
    elif 'soltero' in v or v == 'estado civil':
        if d.get('estado_civil', '').lower().startswith('solt'):
            for cc in coords: txt(*pt(cc), 'X', 9)

    # ── Sección F: firma ──────────────────────────────────────
    elif 'firma' in v:
        for cc in coords: txt(*pt(cc), 'X', 14)

    # ── Sección G: automotor ──────────────────────────────────
    elif v == 'dominio':
        for cc in coords: txt(*pt(cc), d['dominio'])
    elif v == 'marca':
        for cc in coords: txt(*pt(cc), d['marca'])
    elif v == 'tipo':
        for cc in coords: txt(*pt(cc), 'COMPLETAR')
    elif v == 'modelo':
        for cc in coords: txt(*pt(cc), d['modelo'])
    elif 'marca motor' in v:
        for cc in coords: txt(*pt(cc), d.get('marca_motor', 'COMPLETAR'))
    elif 'n de motor' in v or 'nº motor' in v or 'motor' in v:
        for cc in coords: txt(*pt(cc), d['motor'])
    elif 'marca chasis' in v:
        for cc in coords: txt(*pt(cc), d.get('marca_chasis', 'COMPLETAR'))
    elif 'n de chasis' in v or 'nº chasis' in v or 'chasis' in v:
        for cc in coords: txt(*pt(cc), d['chasis'])

    # ── Sección H: solicitud tipo ─────────────────────────────
    elif 'solicitud tipo' in v:
        for cc in coords: txt(*pt(cc), 'X', 9)

    # ── Sección I: modalidades ────────────────────────────────
    elif 'grado' in v:
        for cc in coords: txt(*pt(cc), '1°', 7)
    elif 'clausula de actualizacion' in v or 'clausula' in v:
        for cc in coords: txt(*pt(cc), 'X', 9)
    elif 'concepto' in v:
        for cc in coords: txt(*pt(cc), 'X', 9)

    # ── Recuadros (DNI marcado) ───────────────────────────────
    elif valor == 'RECUADRO' or (not R and len(coords[0]) == 4):
        for cc in coords:
            if len(cc) == 4: box(*cc)

    # ── Fallback: valor literal ───────────────────────────────
    else:
        for cc in coords:
            if len(cc) == 2: txt(*pt(cc), valor)

    return R


def generar_form03(solicitud_path, template_path, output_path, xlsx_path,
                   tipo_op='UVA_PI', carta_path=None):
    if tipo_op not in TIPO_COL:
        print(f"⚠ Tipo '{tipo_op}' no válido. Usando UVA_PI.")
        tipo_op = 'UVA_PI'

    print(f"Tipo de operación : {tipo_op}")
    print(f"Config Excel      : {xlsx_path}")
    print("Parseando solicitud...")
    d = parsear_solicitud(solicitud_path)
    d['tipo_op'] = tipo_op

    if carta_path:
        print("Parseando carta de aprobación...")
        from generar_prenda_autify_v7_working import parsear_carta_aprobacion
        carta = parsear_carta_aprobacion(carta_path)
        d.update(carta)
        # Fallback: si el parser no extrajo del PDF de solicitud (UVA), tomar de carta
        for campo, clave in [('marca','carta_marca'), ('modelo','carta_modelo'),
                              ('anio','carta_anio'), ('categoria','carta_categoria'),
                              ('uso','carta_uso')]:
            if not d.get(campo) and d.get(clave):
                d[campo] = d[clave]

    # Fallback chasis/motor desde solicitud con parser secundario si quedaron vacíos
    if not d.get('chasis') or not d.get('motor'):
        import pdfplumber, re as _re
        try:
            with pdfplumber.open(solicitud_path) as pdf:
                t = ' '.join(p.extract_text() or '' for p in pdf.pages[:3])
                t = _re.sub(r'\s+', ' ', t)
            if not d.get('chasis'):
                m = _re.search(r'N[°º]\s*Chasis[:\s]+(\S+)', t, _re.I)
                if m: d['chasis'] = m.group(1).strip()
            if not d.get('motor'):
                m = _re.search(r'N[°º]\s*Motor[:\s]+(\S+)', t, _re.I)
                if m: d['motor'] = m.group(1).strip()
            if not d.get('marca'):
                m = _re.search(r'Marca[:\s]+([A-Z]{2,15})', t)
                if m: d['marca'] = m.group(1).strip()
            if not d.get('modelo'):
                m = _re.search(r'Modelo[:\s]+([^\n]{5,40}?)(?:\s+Valor|\s+Tipo)', t)
                if m: d['modelo'] = _re.sub(r'\s+',' ', m.group(1)).strip()
        except Exception:
            pass
    if not d.get('chasis'): d['chasis'] = 'COMPLETAR'
    if not d.get('motor'):  d['motor']  = 'COMPLETAR'

    print("\nDatos clave:")
    for k in ('nombre_completo','dni','cuit','dominio','marca','modelo','motor','chasis',
               'monto_insc','localidad','provincia','dom_calle','dom_num','fecha_nac'):
        print(f"  {k:<20}: {d.get(k,'')}")

    entradas = cargar_config_03(xlsx_path, tipo_op)
    print(f"\nCampos activos para {tipo_op}: {len(entradas)}")

    # Agregar recuadros DNI (filas 33 y 38 del Excel — coords sueltas sin variable)
    entradas.append({'coords_str': '50.50-80.54', 'variable': 'Recuadro DNI 1', 'valor': 'RECUADRO'})
    entradas.append({'coords_str': '60.60-80.64', 'variable': 'Recuadro DNI 2', 'valor': 'RECUADRO'})

    reader = PdfReader(template_path)
    writer = PdfWriter()

    for i, page in enumerate(reader.pages):
        packet = io.BytesIO()
        cv = canvas.Canvas(packet, pagesize=(W, H))

        # Form03 es 1 hoja — solo procesamos hoja 1
        if i == 0:
            for e in entradas:
                coords = parse_c(e['coords_str'])
                comandos = resolver_03(e['variable'], e['valor'], coords, d)
                for cmd in comandos:
                    if cmd[0] == 'text':
                        _, col, row, texto, *rest = cmd
                        sz = rest[0] if rest else 6.5
                        draw(cv, col, row, str(texto), sz)
                    elif cmd[0] == 'box':
                        _, c1,r1,c2,r2 = cmd
                        draw_box(cv, c1,r1,c2,r2, fill_color=red, alpha=0.12)

        cv.showPage(); cv.save(); packet.seek(0)
        page.merge_page(PdfReader(packet).pages[0])
        writer.add_page(page)

    with open(output_path, 'wb') as f:
        writer.write(f)
    print(f"\n✓ Form03 generado: {output_path}")


if __name__ == '__main__':
    sol      = sys.argv[1] if len(sys.argv)>1 else None
    template = sys.argv[2] if len(sys.argv)>2 else '/home/claude/form03_grilla.pdf'
    out      = sys.argv[3] if len(sys.argv)>3 else '/home/claude/work/form03_out.pdf'
    xlsx     = sys.argv[4] if len(sys.argv)>4 else '/home/claude/parametros_contrato_prenda_v11.xlsx'
    tipo_op  = sys.argv[5] if len(sys.argv)>5 else 'UVA_PI'
    carta    = sys.argv[6] if len(sys.argv)>6 else None

    if not sol:
        print("Uso: generar_form03_autify.py <solicitud.pdf> <template03.pdf> "
              "<output.pdf> <xlsx> [tipo] [carta.pdf]")
        sys.exit(1)

    generar_form03(sol, template, out, xlsx, tipo_op, carta)
